import os
import re
import pandas as pd
from openpyxl import Workbook

def extract_info(text):
    data = {}
    
    # Diviser le texte en lignes
    lines = text.splitlines()

    # Initialiser les champs à None
    data["Title"] = None
    data["Content"] = None

   # Parcourir les lignes pour extraire les champs
    for i, line in enumerate(lines):
        if not data["Title"]:
            # Supprimer le préfixe "Title:" du texte
            data["Title"] = line.replace("Title:", "").strip()
        
        color_match = re.search(r"Content\s*:\s*([^;]+)", line)
        if color_match:
            # Supprimer le préfixe "Content:" du texte
            data["Content"] = color_match.group(1).strip()

        # Sortir de la boucle si tous les champs sont extraits
        if all(data.values()):
            break

    return data



# Dossier contenant les fichiers d'entrée
dossier_entree = "./"

# Créer un DataFrame vide pour stocker les données de tous les fichiers
all_data = pd.DataFrame()

# Parcourir tous les fichiers du dossier d'entrée
for filename in os.listdir(dossier_entree):
    if filename.endswith(".txt"):
        # Charger le fichier contenant les données
        with open(os.path.join(dossier_entree, filename), "r", encoding="UTF-8") as file:
            data_text = file.read()

        # Extraire les informations du texte
        data = extract_info(data_text)

        # Créer un DataFrame pandas à partir des données
        df = pd.DataFrame([data])

        # Ajouter le DataFrame au DataFrame global
        all_data = pd.concat([all_data, df], ignore_index=True)

# Créer un nouveau fichier Excel avec toutes les données
wb = Workbook()
ws = wb.active

# Écrire les en-têtes dans le fichier Excel
headers = list(all_data.columns)
for c_idx, header in enumerate(headers, start=1):
    ws.cell(row=1, column=c_idx, value=header)

# Écrire les données dans le fichier Excel
for r_idx, row in enumerate(all_data.itertuples(), start=2):  # Commencer à partir de la ligne 2
    for c_idx, value in enumerate(row[1:], start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# Sauvegarder le fichier Excel
wb.save("./dataEnd.xlsx")

print("Fichier Excel avec toutes les données créé avec succès.")
