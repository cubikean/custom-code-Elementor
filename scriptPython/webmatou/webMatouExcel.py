import os
import re
import pandas as pd
from openpyxl import Workbook

def extract_info(text):
    data = {}
    
    # Diviser le texte en lignes
    lines = text.splitlines()

    # Initialiser les champs à None
    data["Nom"] = None
    data["Couleur"] = None
    data["Poids"] = None
    data["Prix"] = None
    data["Ref."] = None
    data["Gamme"] = None
    data["Description"] = None
    data["Dimensions"] = None

    # Variables pour suivre les champs en cours de recherche
    searching_description = False
    searching_dimensions = False

    # Parcourir les lignes pour extraire les champs
    for i, line in enumerate(lines):
        if not data["Nom"]:
            data["Nom"] = line.strip()
        
        color_match = re.search(r"Couleur\s*:\s*([^;]+)", line)
        if color_match:
            data["Couleur"] = color_match.group(1).strip()
        
        poids_match = re.search(r"Poids\s*:\s*([^;]+)", line)
        if poids_match:
            data["Poids"] = poids_match.group(1).strip()

        # Utiliser une expression régulière plus complexe pour extraire le prix
        prix_match = re.search(r"Prix\s*:\s*([\d.,]+)[\s€]*", line)
        if prix_match:
            data["Prix"] = prix_match.group(1).replace(',', '.').strip()

        ref_match = re.search(r"Ref\.\s*:\s*([^;]+)", line)
        if ref_match:
            data["Ref."] = ref_match.group(1).strip()

        gamme_match = re.search(r"Gamme\s*:\s*([^;]+)", line)
        if gamme_match:
            data["Gamme"] = gamme_match.group(1).strip()

        # Rechercher la ligne suivante pour la description
        if searching_description:
            if line.strip():  # Vérifier si la ligne n'est pas vide
                if "Dimensions" not in line:  # Vérifier si on a atteint la section "Dimensions"
                    if not data["Description"]:
                        data["Description"] = line.strip()
                    else:
                        data["Description"] += " " + line.strip()  # Concaténer les lignes de description
                else:
                    searching_description = False

        if "Description" in line:
            searching_description = True

        # Rechercher la ligne suivante pour les dimensions
        if searching_dimensions:
            if line.strip():  # Vérifier si la ligne n'est pas vide
                if "Description" not in line:  # Vérifier si on a atteint la section "Description"
                    if not data["Dimensions"]:
                        data["Dimensions"] = line.strip()
                    else:
                        data["Dimensions"] += " " + line.strip()  # Concaténer les lignes de dimensions
                else:
                    searching_dimensions = False

        if "Dimensions" in line:
            searching_dimensions = True

        # Sortir de la boucle si tous les champs sont extraits
        if all(data.values()):
            break

    return data


# Dossier contenant les fichiers d'entrée
dossier_entree = "C:/Users/Beekom/Desktop/dev/DEV/scriptPython/webmatou/webMatouEndText"

# Créer un DataFrame vide pour stocker les données de tous les fichiers
all_data = pd.DataFrame()

# Parcourir tous les fichiers du dossier d'entrée
for filename in os.listdir(dossier_entree):
    if filename.endswith(".txt"):
        # Charger le fichier contenant les données
        with open(os.path.join(dossier_entree, filename), "r", encoding="latin-1") as file:
            data_text = file.read()

        # Extraire les informations du texte
        data = extract_info(data_text)

        # Créer un DataFrame pandas à partir des données
        df = pd.DataFrame([data])

        # Ajouter le DataFrame au DataFrame global
        all_data = pd.concat([all_data, df], ignore_index=True)

# Créer un nouveau fichier Excel avec toutes les données
wb = Workbook()
ws = wb.active

# Écrire les en-têtes dans le fichier Excel
headers = list(all_data.columns)
for c_idx, header in enumerate(headers, start=1):
    ws.cell(row=1, column=c_idx, value=header)

# Écrire les données dans le fichier Excel
for r_idx, row in enumerate(all_data.itertuples(), start=2):  # Commencer à partir de la ligne 2
    for c_idx, value in enumerate(row[1:], start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# Sauvegarder le fichier Excel
wb.save("C:/Users/Beekom/Desktop/dev/DEV/scriptPython/webmatou/dataEnd.xlsx")

print("Fichier Excel avec toutes les données créé avec succès.")
